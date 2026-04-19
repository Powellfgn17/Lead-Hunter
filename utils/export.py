"""
Export utilities — Convert leads JSON to CSV and Excel formats.
"""

import csv
import json
from pathlib import Path
from typing import Optional

from utils.logger import get_logger

log = get_logger("export")

# Fields to export (ordered for readability)
EXPORT_FIELDS = [
    "name", "nom",
    "address", "adresse",
    "phone", "telephone",
    "email",
    "city", "ville",
    "niche",
    "score",
    "rating",
    "nb_avis",
    "dernier_avis",
    "maps_url", "url_maps",
    "has_website",
    "website_url",
    "years_active",
    "_run_city",
    "_run_niche",
    "_run_timestamp",
]


def _normalize_lead(lead: dict) -> dict:
    """Normalize a lead dict to use consistent field names."""
    return {
        "nom": lead.get("nom", lead.get("name", "")),
        "adresse": lead.get("adresse", lead.get("address", "")),
        "telephone": lead.get("telephone", lead.get("phone", "")),
        "email": lead.get("email", ""),
        "ville": lead.get("ville", lead.get("city", "")),
        "niche": lead.get("niche", ""),
        "score": lead.get("score", ""),
        "rating": lead.get("rating", ""),
        "nb_avis": lead.get("nb_avis", 0),
        "dernier_avis": lead.get("dernier_avis", ""),
        "url_maps": lead.get("url_maps", lead.get("maps_url", "")),
        "has_website": lead.get("has_website", False),
        "website_url": lead.get("website_url", ""),
        "years_active": lead.get("years_active", ""),
        "reseaux_sociaux": _format_socials(lead.get("reseaux_sociaux", {})),
        "run_city": lead.get("_run_city", ""),
        "run_niche": lead.get("_run_niche", ""),
        "run_timestamp": lead.get("_run_timestamp", ""),
    }


def _format_socials(socials: dict) -> str:
    """Format social media dict to a readable string."""
    if not socials:
        return ""
    return " | ".join(f"{k}: {v}" for k, v in socials.items())


# Ordered columns for export
CSV_COLUMNS = [
    "nom", "adresse", "telephone", "email", "ville", "niche",
    "score", "rating", "nb_avis", "dernier_avis", "url_maps",
    "has_website", "reseaux_sociaux", "years_active",
    "run_city", "run_niche", "run_timestamp",
]


def export_csv(leads: list[dict], output_path: Path) -> Path:
    """Export leads to CSV file."""
    normalized = [_normalize_lead(l) for l in leads]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(normalized)

    log.info(f"Exported {len(normalized)} leads to CSV: {output_path}")
    return output_path


def export_excel(leads: list[dict], output_path: Path) -> Path:
    """Export leads to Excel file with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.warning("openpyxl not installed — falling back to CSV export")
        return export_csv(leads, output_path.with_suffix(".csv"))

    normalized = [_normalize_lead(l) for l in leads]
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── Header styling ─────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    headers = {
        "nom": "Nom", "adresse": "Adresse", "telephone": "Téléphone",
        "email": "Email", "ville": "Ville", "niche": "Niche",
        "score": "Score", "rating": "Note", "nb_avis": "Avis",
        "dernier_avis": "Dernier Avis", "url_maps": "Google Maps",
        "has_website": "Site Web ?", "reseaux_sociaux": "Réseaux Sociaux",
        "years_active": "Années Actif",
    }

    display_cols = list(headers.keys())

    for col_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=headers.get(col_name, col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──────────────────────────────────────────
    score_high = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    score_mid = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    score_low = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, lead in enumerate(normalized, 2):
        for col_idx, col_name in enumerate(display_cols, 1):
            value = lead.get(col_name, "")
            # Convert booleans to readable strings
            if isinstance(value, bool):
                value = "Oui" if value else "Non"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Color score column
        score_val = lead.get("score", 0)
        score_cell = ws.cell(row=row_idx, column=display_cols.index("score") + 1)
        if isinstance(score_val, (int, float)):
            if score_val >= 8:
                score_cell.fill = score_high
            elif score_val >= 5:
                score_cell.fill = score_mid
            else:
                score_cell.fill = score_low
        score_cell.font = Font(bold=True, size=12)
        score_cell.alignment = Alignment(horizontal="center")

    # ── Column widths ──────────────────────────────────────
    col_widths = {
        "nom": 25, "adresse": 35, "telephone": 18, "email": 25,
        "ville": 18, "niche": 15, "score": 8, "rating": 8,
        "nb_avis": 8, "dernier_avis": 15, "url_maps": 40,
        "has_website": 12, "reseaux_sociaux": 45, "years_active": 12,
    }
    for col_idx, col_name in enumerate(display_cols, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = col_widths.get(col_name, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{chr(64 + len(display_cols))}1"

    wb.save(output_path)
    log.info(f"Exported {len(normalized)} leads to Excel: {output_path}")
    return output_path


def load_leads_from_json(json_path: Path) -> list[dict]:
    """Load leads from a JSON file."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Handle both flat leads array and run summary format
    if isinstance(data, list):
        if data and "leads" in data[0]:
            # Run summary format — extract leads from each run
            all_leads = []
            for run in data:
                all_leads.extend(run.get("leads", []))
            return all_leads
        return data  # Already flat leads array
    return []
